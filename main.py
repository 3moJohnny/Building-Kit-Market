from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook

window = Tk()          #object because we used class for variable
window.geometry('1100x650')
window.iconbitmap('tools/icon.ico')
window.title('Market Building Tools [متجر أدوات بناء]')

now =datetime.datetime.now()
date = now.strftime("%Y-%m-%d")

#----------> data Base <--------

wb = Workbook()
ws = wb.active

ws.title = 'customer' 
ws["A1"]= 'Full Name'
ws["B1"]= 'Phone Number'
ws["C1"]= 'Address'
ws["D1"]= 'Total'
ws["E1"]= 'Date'
wb.save('johnny.xlsx')


import openpyxl

def save():
    name = enter_name.get()
    phone = enter_phone.get()
    address = enter_address.get()
    total = enter_total.get()
    date_buy = enter_date.get()
    
    excel = openpyxl.load_workbook('johnny.xlsx')
    file = excel.active

    # يُفضل استخدام file.max_row + 1 في كل خط للحفاظ على تسلسل الصفوف
    file.cell(column=1, row=file.max_row + 1, value=name)
    file.cell(column=2, row=file.max_row, value=phone)
    file.cell(column=3, row=file.max_row, value=address)
    file.cell(column=4, row=file.max_row, value=total)
    file.cell(column=5, row=file.max_row, value=date_buy)
    
    excel.save('johnny.xlsx')




#----------> Price <--------

menu = {
    0:['شاكوش', 60],
    1:['دلو', 65],
    2:['قبعة أمان', 65],
    3:['مشرط', 15],
    4:['منشار', 160],
    5:['عربة', 220],
    6:['فأس', 80],
    7:['جاروف', 80],
    8:['بنسه', 40],
    9:['كماشه', 50],
    10:['مفك', 30],
    11:['دريل', 350],
    12:['مسامير', 10],
    13:['منشار كهربي', 450],
    14:['سلم', 90],
    15:['شريط قياس', 35],
    

}


def bill():                                       #  لاضافه المشترايات للقائمه
    global enter_name                             #global make the variable public and we can use it out the function.
    global enter_address
    global enter_total
    global enter_phone
    global enter_date


    lb_image.place(x=1100 , y=530)                 #لاظهار اللوجو
    window.geometry('1400x650')
    F4 = Frame(window, bg='#3F0071', width=296, height=540 , bd=2 , relief=GROOVE)
    F4.place(x=1102,y=1)
    label_name = Label(F4, text='اسم المشتري',  bg='#3F0071' , fg='white'  )   #fg=font color
    label_name.place (x=205,y=10)
    enter_name =Entry(F4,width=24,font=('Tajawal',12),justify=CENTER) #justify هو مكان الادخال
    enter_name.place(x=65,y=45)
    

    label_phone = Label(F4, text='رقم المشتري',  bg='#3F0071' , fg='white'  )   #fg=font color
    label_phone.place (x=207,y=75)
    enter_phone =Entry(F4,width=24,font=('Tajawal',12),justify=CENTER) #justify هو مكان الادخال
    enter_phone.place(x=65,y=105)

    label_address = Label(F4, text='عنوان المشتري',  bg='#3F0071' , fg='white'  )   #fg=font color
    label_address.place (x=209,y=140)
    enter_address =Entry(F4,width=24,font=('Tajawal',12),justify=CENTER) #justify هو مكان الادخال
    enter_address.place(x=65,y=165)

    label_total = Label(F4, text='الحساب الكلي',  bg='#3F0071' , fg='white'  )   #fg=font color
    label_total.place (x=211,y=205)
    enter_total =Entry(F4,width=24,font=('Tajawal',12),justify=CENTER) #justify هو مكان الادخال
    enter_total.place(x=65,y=225)

    label_date = Label(F4, text='التاريخ',  bg='#3F0071' , fg='white'  )   #fg=font color
    label_date.place (x=250,y=270)
    enter_date =Entry(F4,width=24,font=('Tajawal',12),justify=CENTER) #justify هو مكان الادخال
    enter_date.place(x=65,y=290)

    add_button = Button(F4, text='حفظ الفاتورة', width=31, cursor='hand2',bg='#332FD0', command=save)
    add_button.place(x=62 , y=340)

    add_button = Button(F4, text='إفراغ الحقول', width=31, cursor='hand2',bg='#332FD0', command=clear1)
    add_button.place(x=62 , y=390)

    add_button = Button(F4, text='بحث عن مشتري', width=31, cursor='hand2',bg='#332FD0')
    add_button.place(x=62 , y=430)

    add_button = Button(F4, text='حذف فاتورة', width=31, cursor='hand2',bg='#332FD0')
    add_button.place(x=62 , y=470)


    total =0
    for i in trv.get_children():
        trv.delete(i)
    for x in range (len(sb)):                      #range start from sb1 to sb16 
        if (int (sb[x].get())>0):                 #اذا كان عدد المنتج اكبر من صفر ضيفه للسباين بوكس
            price = int(sb[x].get() ) *menu[x] [1]
            total = total + price
            myst = (str(menu [x] [1]) , str(sb[x].get()), str(price))
            trv.insert("",'end',iid=x, text= menu[x] [0], values=myst)
    final = total
    enter_total.insert('1',str(final) + '$')
    enter_date.insert('1',str(date))



def clear():
    for item in trv.get_children():
        trv.delete(item)
    enter_name.delete('0', END)
    enter_address.delete('0', END)
    enter_date.delete('0', END)
    enter_phone.delete('0', END)
    enter_total.delete('0', END)


def clear1():
    enter_name.delete('0', END)
    enter_address.delete('0', END)
    enter_date.delete('0', END)
    enter_phone.delete('0', END)
    enter_total.delete('0', END)


#--------->[Frame1]<------------

F1 = Frame(window, bg='silver', width=650, height=650)
F1.place(x=1,y=1)

#--------->[image]<-----------
img_menu1= PhotoImage(file='tools/1.png')
img_menu2= PhotoImage(file='tools/2.png')
img_menu3= PhotoImage(file='tools/3.png')
img_menu4= PhotoImage(file='tools/4.png')
img_menu5= PhotoImage(file='tools/5.png')
img_menu6= PhotoImage(file='tools/6.png')
img_menu7= PhotoImage(file='tools/7.png')
img_menu8= PhotoImage(file='tools/8.png')
img_menu9= PhotoImage(file='tools/9.png')
img_menu10= PhotoImage(file='tools/10.png')
img_menu11= PhotoImage(file='tools/11.png')
img_menu12= PhotoImage(file='tools/12.png')
img_menu13= PhotoImage(file='tools/13.png')
img_menu14= PhotoImage(file='tools/14.png')
img_menu15= PhotoImage(file='tools/15.png')
img_menu16= PhotoImage(file='tools/16.png')

title = Label(F1, text='مشروع بيع معدات بناء ' , font=('Tajawal 15'),fg='white',bg='#3F0071' , width=80 , )  #tajawal نوع الخط , 15 size ,fg =color , bg=background_color
title.place(x=0,y=0)

menu1 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu1,text='منشار' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu1 .place(x=30,y=45)
menu2 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu2,text='عربة' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu2 .place(x=170,y=45)
menu3 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu3,text='فأس' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu3 .place(x=310,y=45)
menu4 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu4,text='جاروف' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu4 .place(x=450,y=45)


menu5 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu5,text='شاكوش' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu5 .place(x=30,y=180)
menu6 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu6,text='دلو' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu6 .place(x=170,y=180)
menu7 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu7,text='قبعة أمان' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu7 .place(x=310,y=180)
menu8 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu8,text='مشرط' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu8 .place(x=450,y=180)


menu9 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu9,text='بنسه' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu9 .place(x=30,y=315)
menu10 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu10,text='كماشه' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu10.place(x=170,y=315)
menu11 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu11,text='مفك' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu11 .place(x=310,y=315)
menu12 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu12,text='دريل' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu12 .place(x=450,y=315)


menu13 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu13,text='مسامير' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu13 .place(x=30,y=450)
menu14 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu14,text='منشار كهربي' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu14 .place(x=170,y=450)
menu15 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu15,text='سلم' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu15 .place(x=310,y=450)
menu16 = Button(F1, width=88 ,bg='#FB2576',bd=1,relief=SOLID, cursor= 'hand2', height=85 , image=img_menu16,text='شريط قياس' , compound=TOP)      #bd=boarder (حدود الزر) , relief (شكل الجدود),cursor(لما الماوس ييجي ع الكلمه تعطي انيميشن بسيط), compound=TOP(اظهار الصوره اعلي النص)
menu16 .place(x=450,y=450)


#-------------->variables + count <----------

sb=[]
font1 = ('times',12,'normal')
sv1=IntVar()
sv2=IntVar()
sv3=IntVar()
sv4=IntVar()
sv5=IntVar()
sv6=IntVar()
sv7=IntVar()
sv8=IntVar()
sv9=IntVar()
sv10=IntVar()
sv11=IntVar()
sv12=IntVar()
sv13=IntVar()
sv14=IntVar()
sv15=IntVar()
sv16=IntVar()

sb1 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv1)        #زر اظهار عدد مرات شراء المنتج
sb1.place(x=30 , y=275)
sb.append(sb1)
sb2 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv2)        #زر اظهار عدد مرات شراء المنتج
sb2.place(x=170 , y=275)
sb.append(sb2)
sb3 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv3)        #زر اظهار عدد مرات شراء المنتج
sb3.place(x=310 , y=275)
sb.append(sb3)
sb4 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv4)        #زر اظهار عدد مرات شراء المنتج
sb4.place(x=450 , y=275)
sb.append(sb4)

sb5 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv5)        #زر اظهار عدد مرات شراء المنتج
sb5.place(x=30 , y=140)
sb.append(sb5)
sb6 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv6)        #زر اظهار عدد مرات شراء المنتج
sb6.place(x=170 , y=140)
sb.append(sb6)
sb7 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv7)        #زر اظهار عدد مرات شراء المنتج
sb7.place(x=310 , y=140)
sb.append(sb7)
sb8 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv8)        #زر اظهار عدد مرات شراء المنتج
sb8.place(x=450 , y=140)
sb.append(sb8)

sb9 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv9)        #زر اظهار عدد مرات شراء المنتج
sb9.place(x=30 , y=410)
sb.append(sb9)
sb10 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv10)        #زر اظهار عدد مرات شراء المنتج
sb10.place(x=170 , y=410)
sb.append(sb10)
sb11 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv11)        #زر اظهار عدد مرات شراء المنتج
sb11.place(x=310 , y=410)
sb.append(sb11)
sb12 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv12)        #زر اظهار عدد مرات شراء المنتج
sb12.place(x=450 , y=410)
sb.append(sb12)

sb13 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv13)        #زر اظهار عدد مرات شراء المنتج
sb13.place(x=30 , y=545)
sb.append(sb13)
sb14 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv14)        #زر اظهار عدد مرات شراء المنتج
sb14.place(x=170 , y=545)
sb.append(sb14)
sb15 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv15)        #زر اظهار عدد مرات شراء المنتج
sb15.place(x=310 , y=545)
sb.append(sb15)
sb16 =Spinbox (F1,from_=0,to_=5, font=font1, width=10, textvariable=sv16)        #زر اظهار عدد مرات شراء المنتج
sb16.place(x=450 , y=545)
sb.append(sb16)

#-------------> Button <-----------
b1=Button(F1,text='شراء المعدات', fg='white', font=('Tajawal 12'), width=15 , bg='#3F0071', bd=1, relief=SOLID ,cursor='hand2', height=1 , command=bill)
b1.place(x=30,y=600)

b2=Button(F1,text='فاتورة جديدة', fg='white', font=('Tajawal 12'), width=15 , bg='#3F0071', bd=1, relief=SOLID ,cursor='hand2', height=1 , command=clear)
b2.place(x=170,y=600)

b3=Button(F1,text='إستئجار مواد', fg='white', font=('Tajawal 12'), width=15 , bg='#3F0071', bd=1, relief=SOLID ,cursor='hand2', height=1)
b3.place(x=310,y=600)

b4=Button(F1,text='إغلاق البرنامج', fg='white', font=('Tajawal 12'), width=15 , bg='#3F0071', bd=1, relief=SOLID ,cursor='hand2', height=1)
b4.place(x=450,y=600)


#---------> Frame[2] <----------------

F2= Frame(window, bg='gray', width= 440, height=650)
F2.place(x=660, y=1)
trv = ttk.Treeview(F2, selectmode='brows')    #treeview (مسئول عن انشاء الجدول)
trv.place(x=1 , y=1 ,width=438 , height=650 )


trv["columns"]=('1','2','3')
trv.column("#0", width=80 ,anchor='c')           #اسماء المواد      #index=0
trv.column("1", width=50 ,anchor='c')           #السعر
trv.column("2", width=50 ,anchor='c')           #العدد
trv.column("3", width=60 ,anchor='c')           #الحساب الكلي
trv.heading('#0',text=' النوع', anchor='c')
trv.heading('1',text='السعر', anchor='c')
trv.heading('2',text='العدد', anchor='c')
trv.heading('3',text='الحساب الكلي', anchor='c')

im_logo= PhotoImage(file="tools/logo.png")          #استدعاء اللوجو
lb_image=Label(window,image=im_logo)




window.mainloop()



